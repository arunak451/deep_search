import os
import re
from datetime import datetime
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def sanitize_filename(name):
    """
    Sanitizes the company name to be safe for filenames.
    """
    clean_name = re.sub(r'[^a-zA-Z0-9_\- ]', '', name)
    clean_name = clean_name.replace(' ', '_')
    return clean_name[:30]

def get_export_filename(item_name, extension):
    """
    Generates standard formatted file name: DeepDig_PSG_College_2026-06-26.ext
    """
    date_str = datetime.now().strftime("%Y-%m-%d")
    clean_name = sanitize_filename(item_name)
    return f"DeepDig_{clean_name}_{date_str}.{extension}"

def export_to_text(item_name, report_content):
    """
    Saves the report as a simple .txt file.
    """
    filename = get_export_filename(item_name, "txt")
    try:
        with open(filename, "w", encoding="utf-8") as f:
            f.write(report_content)
        print(f"💾 Report saved successfully to text file: {os.path.abspath(filename)}")
        return filename
    except Exception as e:
        print(f"⚠️ Text export failed: {e}")
        return None

def export_to_excel(item_name, report_content, contact_data, news_data, video_data, raw_data_str):
    """
    Saves the data into a multi-sheet Excel file.
    Sheets: Overview, Contact Details, News, Videos, Raw Data
    """
    filename = get_export_filename(item_name, "xlsx")
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 1. Overview Sheet (AI Summary)
            summary_lines = report_content.split('\n')
            df_summary = pd.DataFrame({"Report Summary": summary_lines})
            df_summary.to_excel(writer, sheet_name="Overview", index=False)
            
            # 2. Contact Details Sheet
            contact_rows = []
            # Add Website & Address
            contact_rows.append({"Type": "Website URL", "Details": contact_data.get("website", "N/A"), "Source": "Google Maps"})
            contact_rows.append({"Type": "Address", "Details": contact_data.get("address", "N/A"), "Source": "Google Maps"})
            contact_rows.append({"Type": "Google Maps Phone", "Details": contact_data.get("phone_number", "N/A"), "Source": "Google Maps"})
            
            # Add Scraped emails
            emails = contact_data.get("scraped_emails", [])
            for email in emails:
                contact_rows.append({"Type": "Email", "Details": email, "Source": "Web Scrape"})
                
            # Add Scraped phones
            phones = contact_data.get("scraped_phones", [])
            for phone in phones:
                contact_rows.append({"Type": "Phone", "Details": phone, "Source": "Web Scrape"})
                
            # Add Scraped socials
            socials = contact_data.get("scraped_socials", [])
            for social in socials:
                contact_rows.append({"Type": "Social Media", "Details": social, "Source": "Web Scrape"})
                
            df_contacts = pd.DataFrame(contact_rows)
            df_contacts.to_excel(writer, sheet_name="Contact Details", index=False)
            
            # 3. News Sheet
            if news_data:
                news_rows = []
                for art in news_data:
                    news_rows.append({
                        "Title": art.get("title", "N/A"),
                        "Source": art.get("source_name", "N/A"),
                        "Published Date": art.get("published_date", "N/A"),
                        "URL": art.get("url", "N/A"),
                        "Description": art.get("description", "N/A")
                    })
                df_news = pd.DataFrame(news_rows)
            else:
                df_news = pd.DataFrame(columns=["Title", "Source", "Published Date", "URL", "Description"])
            df_news.to_excel(writer, sheet_name="News", index=False)
            
            # 4. Videos Sheet
            if video_data:
                video_rows = []
                for vid in video_data:
                    video_rows.append({
                        "Title": vid.get("title", "N/A"),
                        "Channel": vid.get("channel_name", "N/A"),
                        "Published Date": vid.get("published_date", "N/A"),
                        "Views": vid.get("view_count", "N/A"),
                        "Likes": vid.get("like_count", "N/A"),
                        "URL": vid.get("url", "N/A")
                    })
                df_videos = pd.DataFrame(video_rows)
            else:
                df_videos = pd.DataFrame(columns=["Title", "Channel", "Published Date", "Views", "Likes", "URL"])
            df_videos.to_excel(writer, sheet_name="Videos", index=False)
            
            # 5. Raw Data Sheet
            raw_lines = raw_data_str.split('\n')
            df_raw = pd.DataFrame({"Raw Scraped Data": raw_lines})
            df_raw.to_excel(writer, sheet_name="Raw Data", index=False)

        # Apply basic auto-column width formatting using openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                # Keep it within reasonable bounds
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
        wb.save(filename)
        
        print(f"📊 Report saved successfully to Excel file: {os.path.abspath(filename)}")
        return filename
    except Exception as e:
        print(f"⚠️ Excel export failed: {e}")
        return None

def _strip_emojis(text):
    """
    Removes emoji and special Unicode characters that ReportLab's Helvetica font
    cannot render. Keeps standard Latin, numbers, punctuation, and common symbols.
    """
    if not text:
        return text
    # Remove emoji and other non-BMP characters
    emoji_pattern = re.compile(
        "["
        "\U0001F600-\U0001F64F"  # emoticons
        "\U0001F300-\U0001F5FF"  # symbols & pictographs
        "\U0001F680-\U0001F6FF"  # transport & map symbols
        "\U0001F1E0-\U0001F1FF"  # flags
        "\U00002702-\U000027B0"  # dingbats
        "\U000024C2-\U0001F251"  # enclosed characters
        "\U0001f926-\U0001f937"  # additional
        "\U00010000-\U0010ffff"  # supplementary
        "\u200d"                 # zero width joiner
        "\u2640-\u2642"          # gender symbols
        "\u2600-\u2B55"          # misc symbols
        "\u23cf"
        "\u23e9"
        "\u231a"
        "\ufe0f"                 # variation selector
        "\u3030"
        "]+",
        flags=re.UNICODE
    )
    cleaned = emoji_pattern.sub('', text)
    # Also clean any remaining chars that Helvetica can't handle (non-Latin-1)
    # Keep ASCII + Latin-1 Supplement (U+0000 to U+00FF) and some extras
    result = []
    for char in cleaned:
        code = ord(char)
        if code < 0x0100 or char in '–—''""•·…€£¥':
            result.append(char)
        else:
            # Replace with space or skip
            result.append(' ')
    return ''.join(result).strip()


def export_to_pdf(item_name, report_content):
    """
    Saves the report as a beautiful styled PDF.
    Parses Markdown elements in the report content (headings, bold, lists) and 
    maps them to ReportLab Flowables.
    Strips emojis and non-Latin chars since ReportLab Helvetica doesn't support them.
    """
    filename = get_export_filename(item_name, "pdf")
    try:
        # Clean the report content of emojis before processing
        report_content = _strip_emojis(report_content)
        clean_item_name = _strip_emojis(item_name)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            filename,
            pagesize=letter,
            rightMargin=54, leftMargin=54,
            topMargin=54, bottomMargin=54
        )
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Custom color definitions (Corporate Theme)
        primary_color = colors.HexColor("#1A365D")   # Deep navy
        secondary_color = colors.HexColor("#2B6CB0") # Slate blue
        text_color = colors.HexColor("#2D3748")      # Charcoal
        
        # Define clean typography styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=24,
            leading=28,
            textColor=primary_color,
            spaceAfter=15
        )
        
        h1_style = ParagraphStyle(
            'SectionH1',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=primary_color,
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )
        
        h2_style = ParagraphStyle(
            'SubSectionH2',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=11,
            leading=15,
            textColor=secondary_color,
            spaceBefore=8,
            spaceAfter=4,
            keepWithNext=True
        )

        body_style = ParagraphStyle(
            'BodyCharcoal',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13.5,
            textColor=text_color,
            spaceAfter=6
        )

        bullet_style = ParagraphStyle(
            'ListBulletCharcoal',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9.5,
            leading=13.5,
            textColor=text_color,
            leftIndent=15,
            firstLineIndent=-10,
            spaceAfter=4
        )
        
        story = []
        
        # Add main report title
        story.append(Paragraph(f"DeepDig Research Report: {clean_item_name}", title_style))
        story.append(Spacer(1, 10))
        
        # Parse Markdown format line by line
        lines = report_content.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                # Add spacing for blank lines
                story.append(Spacer(1, 4))
                continue
            
            # Escape XML special chars for ReportLab (& < >)
            line = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            
            # Replace Markdown Bold **text** with ReportLab bold tags <b>text</b>
            line = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', line)
            
            # Identify headings
            if line.startswith("### "):
                clean_line = line.replace("### ", "", 1)
                story.append(Paragraph(clean_line, h2_style))
            elif line.startswith("## ") or line.startswith("1. ") or line.startswith("2. ") or line.startswith("3. ") or line.startswith("4. ") or line.startswith("5. ") or line.startswith("6. ") or line.startswith("7. ") or line.startswith("8. ") or line.startswith("9. "):
                clean_line = line
                if line.startswith("## "):
                    clean_line = line.replace("## ", "", 1)
                story.append(Paragraph(clean_line, h1_style))
            elif line.startswith("# "):
                clean_line = line.replace("# ", "", 1)
                story.append(Paragraph(clean_line, title_style))
            # Identify bullets
            elif line.startswith("* ") or line.startswith("- "):
                clean_line = line[2:]
                story.append(Paragraph(f"&bull; {clean_line}", bullet_style))
            else:
                story.append(Paragraph(line, body_style))
                
        # Build PDF
        doc.build(story)
        print(f"PDF Report saved successfully: {os.path.abspath(filename)}")
        return filename
    except Exception as e:
        print(f"⚠️ PDF export failed: {e}")
        return None

